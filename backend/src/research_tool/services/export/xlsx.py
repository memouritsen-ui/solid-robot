"""XLSX export provider using openpyxl."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .exporter import Exporter, ExportFormat, ExportResult, ResearchExportData


class XLSXExporter(Exporter):
    """Export research results to XLSX format using openpyxl."""

    @property
    def format(self) -> ExportFormat:
        """Return the export format."""
        return ExportFormat.XLSX

    @property
    def mime_type(self) -> str:
        """Return the MIME type."""
        return (
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        )

    @property
    def file_extension(self) -> str:
        """Return the file extension."""
        return "xlsx"

    async def export(self, data: ResearchExportData) -> ExportResult:
        """Export research data to XLSX.

        Args:
            data: Research data to export

        Returns:
            ExportResult: XLSX content as bytes
        """
        try:
            wb = self._generate_workbook(data)
            buffer = BytesIO()
            wb.save(buffer)
            xlsx_bytes = buffer.getvalue()

            return ExportResult(
                success=True,
                format=self.format,
                content=xlsx_bytes,
                filename=self.generate_filename(data.query),
                mime_type=self.mime_type,
            )
        except Exception as e:
            return ExportResult(
                success=False,
                format=self.format,
                content=None,
                filename=self.generate_filename(data.query),
                mime_type=self.mime_type,
                error=str(e),
            )

    def _generate_workbook(self, data: ResearchExportData) -> Workbook:
        """Generate XLSX workbook from research data.

        Args:
            data: Research data

        Returns:
            Workbook: openpyxl Workbook object
        """
        wb = Workbook()

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4A90D9", end_color="4A90D9", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Summary Sheet
        ws = wb.active
        ws.title = "Summary"

        ws["A1"] = "Research Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        ws["A3"] = "Query"
        ws["B3"] = data.query
        ws["A4"] = "Domain"
        ws["B4"] = data.domain
        ws["A5"] = "Confidence Score"
        ws["B5"] = f"{data.confidence_score:.1%}"
        ws["A6"] = "Generated"
        ws["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        ws["A8"] = "Executive Summary"
        ws["A8"].font = Font(bold=True)
        ws["A9"] = data.summary
        ws.merge_cells("A9:D9")

        # Style column A as labels
        for row in range(3, 7):
            ws.cell(row=row, column=1).font = Font(bold=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60

        # Facts Sheet
        ws_facts = wb.create_sheet("Findings")

        # Headers
        headers = ["#", "Finding", "Confidence", "Source", "Verified"]
        for col, header in enumerate(headers, 1):
            cell = ws_facts.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for i, fact in enumerate(data.facts, 1):
            row = i + 1
            ws_facts.cell(row=row, column=1, value=i).border = thin_border
            ws_facts.cell(
                row=row,
                column=2,
                value=fact.get("statement", fact.get("content", "")),
            ).border = thin_border
            ws_facts.cell(
                row=row, column=3, value=f"{fact.get('confidence', 0.0):.1%}"
            ).border = thin_border
            ws_facts.cell(
                row=row, column=4, value=fact.get("source", "Unknown")
            ).border = thin_border
            ws_facts.cell(
                row=row, column=5, value="Yes" if fact.get("verified") else "No"
            ).border = thin_border

        # Column widths
        ws_facts.column_dimensions["A"].width = 5
        ws_facts.column_dimensions["B"].width = 60
        ws_facts.column_dimensions["C"].width = 12
        ws_facts.column_dimensions["D"].width = 25
        ws_facts.column_dimensions["E"].width = 10

        # Sources Sheet
        ws_sources = wb.create_sheet("Sources")

        headers = ["#", "Title", "URL", "Type", "Reliability"]
        for col, header in enumerate(headers, 1):
            cell = ws_sources.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for i, source in enumerate(data.sources, 1):
            row = i + 1
            ws_sources.cell(row=row, column=1, value=i).border = thin_border
            ws_sources.cell(
                row=row, column=2, value=source.get("title", "Untitled")
            ).border = thin_border
            ws_sources.cell(
                row=row, column=3, value=source.get("url", "")
            ).border = thin_border
            ws_sources.cell(
                row=row, column=4, value=source.get("type", "web")
            ).border = thin_border
            ws_sources.cell(
                row=row,
                column=5,
                value=f"{source.get('reliability_score', 0.0):.1%}",
            ).border = thin_border

        ws_sources.column_dimensions["A"].width = 5
        ws_sources.column_dimensions["B"].width = 40
        ws_sources.column_dimensions["C"].width = 50
        ws_sources.column_dimensions["D"].width = 15
        ws_sources.column_dimensions["E"].width = 12

        # Limitations Sheet
        ws_limits = wb.create_sheet("Limitations")

        ws_limits["A1"] = "Research Limitations"
        ws_limits["A1"].font = Font(bold=True, size=12)

        for i, limitation in enumerate(data.limitations, 1):
            ws_limits.cell(row=i + 2, column=1, value=f"{i}.")
            ws_limits.cell(row=i + 2, column=2, value=limitation)

        ws_limits.column_dimensions["A"].width = 5
        ws_limits.column_dimensions["B"].width = 80

        return wb
